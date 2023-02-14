###############################################################################
# プレイリストから再生数、コメント数、高評価、低評価を入力する
# ファイルは時刻をくっつけて別名で保存される
###############################################################################

from googleapiclient.discovery import build
import json
import openpyxl
import re
import datetime
from datetime import date
from openpyxl.styles import Font

#================================================================================
# 初期化
#================================================================================
def initYoutube(API_KEY):
    API_SERVICE_NAME = "youtube"
    API_VERSION = "v3"
    return build(API_SERVICE_NAME, API_VERSION, developerKey=API_KEY)

#================================================================================
# プレイリストIDを渡して動画IDリストを得る
#================================================================================
def getIdListFromPlaylist(id_,youtube):

    nextPageToken = 'start'
    response = []

    while(nextPageToken is not None):

        if(nextPageToken == 'start'):
            search_response = youtube.playlistItems().list(
            part= 'snippet',
            playlistId=id_,
            maxResults = 50,
            ).execute()
        else:
            search_response = youtube.playlistItems().list(
            part= 'snippet',
            playlistId=id_,
            maxResults = 50,
            pageToken = nextPageToken
            ).execute()

        if('nextPageToken' in search_response):
            nextPageToken = search_response['nextPageToken']
        else:
            nextPageToken = None
        
        for item in search_response['items']:
            response.append(item['snippet']['resourceId']['videoId'])

 
    response.reverse()   
    return response

#================================================================================
# YoutubeのAPIを叩いて統計情報を取得する
#================================================================================
def getCountDetails(id_, youtube):

    #50件ずつに分割
    idLists = split_list(id_,50)
    response = []

    for idList in idLists:
        search_response = youtube.videos().list(
        part= 'statistics,snippet',
        id=idList,
        ).execute()

        response.extend(search_response['items'])

    return response

#================================================================================
# 指定ワークシートからIDリストを取得して数値を更新する
#================================================================================
def setCountDetail(ws,idList,youtube):

    url_list = []

    result = getCountDetails(idList,youtube)

    row = 1
    ws.cell(row,1).value = '公開日'
    ws.cell(row,2).value = 'タイトル'
    ws.cell(row,3).value = 'URL'     
    ws.cell(row,4).value = '再生数'
    ws.cell(row,5).value = 'コメント数'
    ws.cell(row,6).value = '高評価'
    ws.cell(row,7).value = '低評価'

    row += 1

    for item in result:

        published = datetime.datetime.fromisoformat(item['snippet']['publishedAt'].replace('Z', '+00:00')).strftime('%Y/%m/%d')
        ws.cell(row,1).value = published
        ws.cell(row,2).value = item['snippet']['title']
        ws.cell(row,3).value = 'https://www.youtube.com/watch?v='+item['id']
        
        url_list.append('https://www.youtube.com/watch?v='+item['id'])

        #コメント非公開等あるとそもそも値がないので、その場合は-1にしておく
        ws.cell(row,4).value = int(item['statistics']['viewCount']) if 'viewCount' in item['statistics'] else -1
        ws.cell(row,5).value = int(item['statistics']['commentCount'])  if 'commentCount' in item['statistics'] else -1
        ws.cell(row,6).value = int(item['statistics']['likeCount'])  if 'likeCount' in item['statistics'] else -1
        ws.cell(row,7).value = int(item['statistics']['dislikeCount']) if 'dislikeCount' in item['statistics'] else -1

        #ws.cell(row,3).hyperlink = ws.cell(row,3).value
        ws.cell(row,4).number_format = '#,##0'
        ws.cell(row,5).number_format = '#,##0'
        ws.cell(row,6).number_format = '#,##0'
        ws.cell(row,7).number_format = '#,##0'

        row += 1

#================================================================================
# 配列を指定した個数ごとに分割
#================================================================================
def split_list(l, n):
    for idx in range(0, len(l), n):
        yield l[idx:idx + n]



# YoutubeAPI用キー
API_KEY = 'AIzaSyCdaS1JZZ3IArBYwQShFZ4tjQx_YzrOqZQ' #ここに各々で取得したYoutube用のAPIキーを入れる
FILENAME = 'Youtube'
youtube = initYoutube(API_KEY)

#「ワークシート名:プレイリストID」の辞書型配列にしておく
playList = {
'DJ downloadlist':'PLQ2CDHLBLZQGgUvuwjUNkls5UJpdPbCuf',
}

# Excelファイル新規作成
wb = openpyxl.Workbook()

for key in playList:
    wb.create_sheet(key,0)
    setCountDetail(wb[key],getIdListFromPlaylist(playList[key],youtube),youtube)    

#タイムスタンプ付けて保存
dt_now = datetime.datetime.now()
wb.save(FILENAME+dt_now.strftime('_%Y%m%d_%H%M')+'.xlsx')

